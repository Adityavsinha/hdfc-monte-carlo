"""
dashboard_export.py
-------------------
Pushes simulation results to:
  A) Excel workbook (openpyxl)  — always available, no auth needed
  B) Google Sheets (gspread)    — optional, requires service-account JSON
"""

import os
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

from simulation_engine import SimResults

logger = logging.getLogger(__name__)
OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ════════════════════════════════════════════
#  EXCEL EXPORT
# ════════════════════════════════════════════

_HEADER_FILL = PatternFill("solid", fgColor="0F2537")
_ACCENT_FILL = PatternFill("solid", fgColor="1F4E79")
_GREEN_FILL  = PatternFill("solid", fgColor="1E4620")
_RED_FILL    = PatternFill("solid", fgColor="4A1212")
_GOLD_FILL   = PatternFill("solid", fgColor="3D2B00")

_H_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_N_FONT  = Font(name="Calibri", size=10)
_THIN    = Side(style="thin", color="2E75B6")
_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER  = Alignment(horizontal="center", vertical="center")
_RIGHT   = Alignment(horizontal="right", vertical="center")


def _cell(ws, row, col, value, font=None, fill=None, align=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:    c.font      = font
    if fill:    c.fill      = fill
    if align:   c.alignment = align
    if num_fmt: c.number_format = num_fmt
    c.border = _BORDER
    return c


def _section_header(ws, row, col, title, width=2):
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row,   end_column=col + width - 1
    )
    c = ws.cell(row=row, column=col, value=title)
    c.font      = _H_FONT
    c.fill      = _HEADER_FILL
    c.alignment = _CENTER
    c.border    = _BORDER


def _write_summary_sheet(wb: openpyxl.Workbook, results: SimResults) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    _section_header(ws, 1, 1, f"HDFC Bank — Monte Carlo Report  [{now}]", width=3)

    # ── Price projections ──
    rows = [
        ("PRICE PROJECTIONS", None, None, _ACCENT_FILL),
        ("Current Price (S₀)",             f"₹{results.S0:,.2f}",                 None,                          None),
        (f"Expected Price (Day {results.config.horizon_days})",
                                            f"₹{results.mean_price:,.2f}",
                                            f"{(results.mean_price/results.S0-1)*100:+.2f}%",  _GREEN_FILL if results.mean_price > results.S0 else _RED_FILL),
        ("Median Price",                   f"₹{results.median_price:,.2f}",        None,   None),
        ("5th Percentile",                 f"₹{results.ci_5:,.2f}",               f"{(results.ci_5/results.S0-1)*100:+.2f}%",  _RED_FILL),
        ("95th Percentile",                f"₹{results.ci_95:,.2f}",              f"{(results.ci_95/results.S0-1)*100:+.2f}%", _GREEN_FILL),
        ("RISK METRICS", None, None, _ACCENT_FILL),
        ("VaR (95%)",                      f"₹{results.var_95:,.2f}",             "max 1-day loss at 95% conf",  _RED_FILL),
        ("VaR (99%)",                      f"₹{results.var_99:,.2f}",             "max 1-day loss at 99% conf",  _RED_FILL),
        ("CVaR / Expected Shortfall (95%)",f"₹{results.cvar_95:,.2f}",            "avg loss beyond VaR",         _RED_FILL),
        ("PROBABILITY METRICS", None, None, _ACCENT_FILL),
        ("P(Price Increases)",             f"{results.prob_increase:.2%}",         None, _GREEN_FILL if results.prob_increase > 0.5 else _RED_FILL),
        ("P(Price > +10%)",                f"{results.prob_10pct_up:.2%}",         None, None),
        ("P(Price < -10%)",                f"{results.prob_10pct_down:.2%}",       None, None),
        ("SIMULATION PARAMETERS", None, None, _ACCENT_FILL),
        ("Simulations",                    f"{results.config.n_simulations:,}",    None, None),
        ("Horizon (days)",                 f"{results.config.horizon_days}",       None, None),
        ("Daily Drift (μ)",                f"{results.mu:.6f}",                    None, None),
        ("Daily Volatility (σ)",           f"{results.sigma:.6f}",                 None, None),
        ("Annual Volatility",              f"{results.sigma * (252**0.5) * 100:.2f}%", None, None),
        ("Fat-tail Model",                 "Student-t" if results.config.use_fat_tails else "Normal", None, None),
    ]

    r = 2
    for label, val, note, fill in rows:
        if val is None:   # section header row
            _section_header(ws, r, 1, label, width=3)
        else:
            _cell(ws, r, 1, label, font=_N_FONT, fill=fill, align=_RIGHT)
            _cell(ws, r, 2, val,   font=Font(name="Calibri", bold=True, size=10,
                                             color="92D050" if fill == _GREEN_FILL
                                             else "FF7575" if fill == _RED_FILL else "FFFFFF"),
                  fill=fill, align=_CENTER)
            if note:
                _cell(ws, r, 3, note, font=Font(name="Calibri", size=9, color="9DC3E6"),
                      align=_RIGHT)
        r += 1

    ws.freeze_panes = "A2"


def _write_percentile_sheet(wb: openpyxl.Workbook, results: SimResults) -> None:
    ws = wb.create_sheet("Percentile Paths")
    T  = results.config.horizon_days

    headers = ["Day", "P5 (₹)", "P25 (₹)", "P50 (₹)", "P75 (₹)", "P95 (₹)"]
    for c, h in enumerate(headers, 1):
        _cell(ws, 1, c, h, font=_H_FONT, fill=_HEADER_FILL, align=_CENTER)
        ws.column_dimensions[get_column_letter(c)].width = 14

    p5  = np.percentile(results.paths, 5,  axis=1)
    p25 = np.percentile(results.paths, 25, axis=1)
    p50 = np.percentile(results.paths, 50, axis=1)
    p75 = np.percentile(results.paths, 75, axis=1)
    p95 = np.percentile(results.paths, 95, axis=1)

    for i in range(T):
        r = i + 2
        ws.cell(r, 1, i + 1)
        for c, v in enumerate([p5[i], p25[i], p50[i], p75[i], p95[i]], 2):
            ws.cell(r, c, round(v, 2)).number_format = '₹#,##0.00'

    # Add line chart
    chart = LineChart()
    chart.title  = "Monte Carlo Percentile Paths"
    chart.y_axis.title = "Price (₹)"
    chart.x_axis.title = "Days"
    chart.width  = 30
    chart.height = 18
    chart.style  = 10

    data_ref = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=T + 1)
    chart.add_data(data_ref, titles_from_data=True)
    for i, color in enumerate(["FF4444", "FFAA44", "44AAFF", "44FF88", "44FF44"]):
        chart.series[i].graphicalProperties.line.solidFill = color
        chart.series[i].graphicalProperties.line.width     = 20000

    ws.add_chart(chart, "H2")
    ws.freeze_panes = "A2"


def _write_histogram_sheet(wb: openpyxl.Workbook, results: SimResults) -> None:
    ws = wb.create_sheet("Terminal Distribution")
    final = results.final_prices

    # Build histogram bins
    bins    = np.linspace(final.min(), final.max(), 60)
    counts, edges = np.histogram(final, bins=bins)
    centers = (edges[:-1] + edges[1:]) / 2

    ws.cell(1, 1, "Bin Centre (₹)").font = _H_FONT
    ws.cell(1, 2, "Frequency").font      = _H_FONT
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    for i, (c, v) in enumerate(zip(centers, counts), 2):
        ws.cell(i, 1, round(c, 2)).number_format = '₹#,##0.00'
        ws.cell(i, 2, int(v))

    chart = BarChart()
    chart.title       = f"Terminal Price Distribution  (Day {results.config.horizon_days})"
    chart.y_axis.title = "Frequency"
    chart.x_axis.title = "Terminal Price"
    chart.width  = 30
    chart.height = 18
    chart.style  = 10
    chart.grouping = "clustered"

    data_ref = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=len(counts) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.series[0].graphicalProperties.solidFill = "1F77B4"
    ws.add_chart(chart, "D2")


def export_to_excel(
    results     : SimResults,
    log_returns : pd.Series,
    filename    : str = "hdfc_monte_carlo.xlsx",
) -> Path:
    path = OUTPUT_DIR / filename
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    _write_summary_sheet(wb, results)
    _write_percentile_sheet(wb, results)
    _write_histogram_sheet(wb, results)

    wb.save(path)
    logger.info(f"Excel dashboard saved → {path}")
    return path


# ════════════════════════════════════════════
#  GOOGLE SHEETS EXPORT  (optional)
# ════════════════════════════════════════════

def export_to_gsheets(
    results          : SimResults,
    spreadsheet_id   : str,
    creds_json_path  : str = "service_account.json",
) -> None:
    """
    Requires:
      pip install gspread google-auth
      A GCP service account with Sheets API enabled.
      Share the spreadsheet with the service-account email.
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        logger.error("Run:  pip install gspread google-auth")
        return

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(creds_json_path, scopes=scopes)
    gc    = gspread.authorize(creds)
    sh    = gc.open_by_key(spreadsheet_id)

    # ── Summary tab ──
    try:
        ws = sh.worksheet("Summary")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet("Summary", rows=50, cols=5)

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    data = [
        ["HDFC Bank Monte Carlo", now],
        [""],
        ["Metric", "Value", "Change vs Current"],
        ["Current Price",         results.S0,          "—"],
        ["Expected Price",        round(results.mean_price, 2),
                                  f"{(results.mean_price/results.S0-1)*100:+.2f}%"],
        ["Median Price",          round(results.median_price, 2), ""],
        ["5th Percentile",        round(results.ci_5, 2),
                                  f"{(results.ci_5/results.S0-1)*100:+.2f}%"],
        ["95th Percentile",       round(results.ci_95, 2),
                                  f"{(results.ci_95/results.S0-1)*100:+.2f}%"],
        ["VaR 95%",               round(results.var_95, 2), ""],
        ["VaR 99%",               round(results.var_99, 2), ""],
        ["CVaR 95%",              round(results.cvar_95, 2), ""],
        ["P(Price Increases)",    f"{results.prob_increase:.2%}", ""],
        ["P(Price > +10%)",       f"{results.prob_10pct_up:.2%}", ""],
        ["P(Price < -10%)",       f"{results.prob_10pct_down:.2%}", ""],
        ["Annual Volatility",     f"{results.sigma*(252**0.5)*100:.2f}%", ""],
        ["Simulations",           results.config.n_simulations, ""],
        ["Horizon (days)",        results.config.horizon_days, ""],
    ]
    ws.clear()
    ws.update("A1", data)
    logger.info("Google Sheets updated ✓")
